from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "templates"
OUTPUT.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="0B6B72")
NOTE_FILL = PatternFill("solid", fgColor="DFF3F4")
HEADER_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
BODY_FONT = Font(name="Microsoft YaHei", color="173B3F")


def style_header(sheet, row, columns):
    for column in range(1, columns + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_body(sheet, first_row, last_row, columns):
    for row in sheet.iter_rows(min_row=first_row, max_row=last_row, max_col=columns):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.protection = Protection(locked=False)


def student_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "学生导入"
    headers = [
        "xue_hao", "xing_ming", "ban_ji_bian_ma", "nian_ji",
        "yong_hu_ming", "chu_shi_mi_ma", "zhang_hao_zhuang_tai",
    ]
    sheet.append(headers)
    sheet.append(["S20260001", "示例学生甲", "CLASS_TEMPLATE", "高三", "S20260001", "DemoPass1", "ENABLED"])
    sheet.append(["S20260002", "示例学生乙", "CLASS_TEMPLATE", "高三", "", "", "ENABLED"])
    style_header(sheet, 1, len(headers))
    style_body(sheet, 2, 501, len(headers))
    widths = [18, 18, 22, 12, 20, 20, 22]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:G501"
    status = DataValidation(type="list", formula1='"ENABLED,DISABLED"', allow_blank=True)
    sheet.add_data_validation(status)
    status.add("G2:G501")

    guide = workbook.create_sheet("填写说明")
    guide.append(["字段", "必填", "约束与处理"])
    rows = [
        ("xue_hao", "是", "1–64 字符，全局唯一；按文本保存可保留前导零。"),
        ("xing_ming", "是", "2–32 字符；模板只使用虚构姓名。"),
        ("ban_ji_bian_ma", "是", "必须替换为系统中已存在且 ACTIVE 的班级编码。"),
        ("nian_ji", "是", "必须与班级年级完全一致。"),
        ("yong_hu_ming", "否", "为空时使用学号；仅字母、数字、点、下划线、连字符，1–64 字符。"),
        ("chu_shi_mi_ma", "否", "8–64 位且同时含字母和数字；为空则确认导入时随机生成。"),
        ("zhang_hao_zhuang_tai", "否", "ENABLED 或 DISABLED；为空默认 ENABLED。"),
        ("处理流程", "", "先 Preview；全部 VALID 后再 Confirm。Confirm 整批事务写入并返回一次性初始密码。"),
        ("文件限制", "", "只接受 .xlsx；最大 5 MB；最多 500 个非空数据行；禁止公式和宏。"),
    ]
    for row in rows:
        guide.append(row)
    style_header(guide, 1, 3)
    style_body(guide, 2, guide.max_row, 3)
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 12
    guide.column_dimensions["C"].width = 92
    guide.freeze_panes = "A2"
    workbook.save(OUTPUT / "student-import-template.xlsx")


def question_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "题目检查"
    sheet.merge_cells("A1:S1")
    sheet["A1"] = "RIKE 题目导入模板（示例为项目自编题；导入后仅进入 PENDING，须人工审核）"
    sheet["A1"].fill = NOTE_FILL
    sheet["A1"].font = Font(name="Microsoft YaHei", bold=True, color="0B4F54", size=12)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    headers = [
        "学科", "年份", "区域", "试卷来源", "题号", "题型", "题干", "选项", "答案", "标准解析",
        "题干图片数", "解析图片数", "一级知识点", "知识点", "难度", "难度说明",
        "答案解析审核状态", "审核状态", "来源文件",
    ]
    sheet.append(headers)
    sheet.append([
        "物理", "2026", "项目自编", "RIKE 自编训练题", "TEMPLATE-001", "单选题",
        "一列简谐横波的频率为 5 Hz，波长为 2 m，其波速是多少？",
        "A. 2.5 m/s\nB. 7 m/s\nC. 10 m/s\nD. 25 m/s", "C",
        "依据波速公式 v=fλ，代入 f=5 Hz、λ=2 m，得到 v=10 m/s。",
        "0", "0", "力学", "力学>机械振动与机械波>波速、波长与频率", "easy",
        "考查波速公式的直接应用。", "待审核", "待审核",
        "试题文件：解析方案.md\n答案解析文件：来源合规检查.md",
    ])
    style_header(sheet, 2, len(headers))
    style_body(sheet, 3, 102, len(headers))
    widths = [10, 10, 14, 22, 16, 14, 48, 38, 18, 48, 14, 14, 18, 46, 12, 28, 20, 14, 42]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index) if index <= 26 else str(index)].width = width
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = "A2:S102"
    type_rule = DataValidation(type="list", formula1='"单选题,多选题,实验填空题,解答题"')
    subject_rule = DataValidation(type="list", formula1='"物理,化学,生物"')
    difficulty_rule = DataValidation(type="list", formula1='"easy,medium,hard"')
    sheet.add_data_validation(type_rule)
    sheet.add_data_validation(subject_rule)
    sheet.add_data_validation(difficulty_rule)
    type_rule.add("F3:F102")
    subject_rule.add("A3:A102")
    difficulty_rule.add("O3:O102")

    guide = workbook.create_sheet("填写说明")
    guide.append(["主题", "说明"])
    notes = [
        ("固定格式", "工作表必须名为“题目检查”，标题在第 1 行，19 列表头在第 2 行且顺序固定；第 3 行起为数据。"),
        ("范围", "同一文件只能包含一个学科；最多 100 行；只接受 .xlsx，最大 10 MB；数据区域禁止公式。"),
        ("题型", "支持单选题、多选题、实验填空题、解答题；选项格式为“标签. 内容”，每项换行。"),
        ("答案", "单选填一个标签；多选可用“、”分隔；填空使用“①. 内容 ②. 内容”；解答题填写参考答案原文。"),
        ("知识点", "“知识点”须填写当前学科已存在的完整路径；多个路径使用换行分隔。"),
        ("难度", "仅 easy、medium、hard，系统映射为 1、3、5。"),
        ("来源", "必须同时声明“试题文件”和“答案解析文件”，相对路径从受控题库目录解析；请替换为有权使用的本地来源。"),
        ("附件", "正文以〔图片对象 I001〕或〔公式对象 F001〕引用；文件名须与题号、类型和序号精确匹配。"),
        ("状态", "Preview 不写库；Confirm 重新校验文件哈希，整批写入 PENDING，必须由人工审核后才可发布。"),
        ("示例权利", "示例题为项目自编，仅用于验证模板，不冒充高考真题、教材原题或官方试题。"),
    ]
    for row in notes:
        guide.append(row)
    style_header(guide, 1, 2)
    style_body(guide, 2, guide.max_row, 2)
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 110
    guide.freeze_panes = "A2"
    workbook.save(OUTPUT / "question-import-template.xlsx")


if __name__ == "__main__":
    student_template()
    question_template()
